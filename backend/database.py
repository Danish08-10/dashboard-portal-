import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
from sqlalchemy import create_engine, inspect
import os
from openpyxl import load_workbook


def get_sheet_by_prefix(file_path, prefix):
    workbook = load_workbook(filename=file_path, read_only=True)
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith(prefix):
            print("Matched sheet:", sheet_name)
            return sheet_name
    print("No sheet found starting with:", prefix)
    return None


user = "postgres.kimsluaqkkipyzpnfpxl"
password = "CxzxdPqFKciVppqT"
host = "aws-1-ap-southeast-2.pooler.supabase.com"
port = "5432"
database = "CBS Central"
table_name = "insale"


sheet=get_sheet_by_prefix(r"C:\Users\danish.azhar\Desktop\Updated Deployment-Tracker.xlsx",prefix='Deployment-Tracker')
df=pd.read_excel(r"C:\Users\danish.azhar\Desktop\Updated Deployment-Tracker.xlsx",sheet_name=sheet)
df_invoices_mrc=pd.read_excel(r"C:\Users\danish.azhar\Desktop\Invoices\MRC & OTC of all Links.xlsx",sheet_name='Sheet1')
df_vendor_status=pd.read_excel(r"C:\Users\danish.azhar\Desktop\Vendor Links Working\Invoices status.xlsx",sheet_name='Sheet1')


try:
    engine = create_engine(
        f"postgresql://{user}:{password}@{host}:{port}/postgres",
        pool_pre_ping=True
    )

    with engine.connect() as conn:
        print("Supabase Engine Connected Successfully")

except Exception as e:
    print(f"Error is this: {e}")

df.to_sql(table_name, con=engine, if_exists="replace", index=False)
df_invoices_mrc.to_sql("MRC", con=engine, if_exists="replace", index=False)
df_vendor_status.to_sql("Vendor_Status", con=engine, if_exists="replace", index=False)

print("Upload successful.")